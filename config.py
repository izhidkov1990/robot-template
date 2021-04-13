import openpyxl

def read(path, mode) -> dict:

    """mode - 'Test' or 'Prod' """

    if mode =='Test':
        sheet_index = 1
    elif mode == 'Prod':
        sheet_index = 0
    else:
        sheet_index = 1

    workbook = openpyxl.load_workbook(path)
    workbook.active = sheet_index
    sheet = workbook.active
    data = {}
    for row_index in range(2, sheet.max_row + 1):  # 1-skip heading row
        config_name = sheet.cell(row_index, 1).value
        config_value = sheet.cell(row_index, 2).value
        data.update({config_name: config_value})  # add data to dictionary
    return data

class Config:
    def __init__(self, dict: dict, mode: str):
        self.dict = dict
        self.mode = mode
