import glob
import os
import openpyxl
import error_handler


"""Получение задач"""

def get() -> list:

    """Эмуляция получения задач"""

    transaction_list = []
    data = ['Transaction1', 2, 'Transaction3']
    for item in data:
        transaction = Transaction(item)
        transaction.add_info1 = 'some info about transaction1'
        # transaction.add_info2 = 'some info about transaction2'
        transaction_list.append(transaction)
    return transaction_list


class Transaction:

    """Класс транзакции
    processed_object - обрабатываемый обьект(файл, письмо, строка из Excel и.т.д....)"""

    def __init__(self, transaction):
        self.processed_object = transaction
        # TODO Может размещаться дополнительная информация о транзакции
        self.add_info1 = None
        self.add_info2 = None


class Miner:

    @staticmethod
    def get_files_data(path: str, file_type: str=None) -> list:

        """Получение списка файлов
        path: - путь к папке
        file_type: - расширение файла"""
        
        if file_type:
            if file_type.startswith('.'):
                return glob.glob(path + r'\\' + '*' + file_type)
            else:
                return glob.glob(path + r'\\' + '*.' + file_type)
        else:
            return glob.glob(path+ r'\\' + '*.*')

    @staticmethod
    def get_rows_data(path: str, sheet: int=None, use_headers: bool=None) -> list or dict:

        """Функция возвращает список из словарей(строк) файла Excel
        path: путь до файла Excel
        use_headers: bool - учитывать заголовки столбцов и возвращать словари
        sheet: int - выбол листа книги. Если не задано - первый лист"""

        if not sheet:
            sheet = 0
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)  # read_only=True ускоряет чтение больших Excel
        workbook.active = sheet
        sheet = workbook.active
        rows_list = []
        row = {}
        if not sheet.max_row > 100:
            if use_headers:
                # Возвратит список rows_list из словарей типа {'column_name': 'cell_value'}
                columns_name = []
                for column_index in range(1, sheet.max_column+1):
                    columns_name.append(sheet.cell(1, column_index).value)
                for row_index in range(2, sheet.max_row+1): 
                    row = {}
                    for name in columns_name:
                        cell_value = sheet.cell(row_index, columns_name.index(name)+1).value
                        row.update({name: cell_value}) 
                    rows_list.append(row)
                return rows_list
            else:
                # Возвратит список из строк со значениями ячеек 
                for row_index in range(1, sheet.max_row): 
                    row = []
                    for column_index in range(1, sheet.max_column+1):               
                        cell_value = sheet.cell(row_index, column_index).value
                        row.append(cell_value)
                    rows_list.append(row)
                return rows_list
        else:
            """Перебор `100 и более строк в цикле заментно замедляет процесс чтения.
            Для чтения больших массивов данных из Excel лучше использовать pandas"""
            # data_frame = pd.DataFrame(sheet.values)
            raise Exception('Too large excel file')
            
